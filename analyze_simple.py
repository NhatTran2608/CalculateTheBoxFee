# -*- coding: utf-8 -*-
import pandas as pd
import openpyxl
from openpyxl import load_workbook

# Đọc file Excel chi tiết để hiểu công thức
file_path = 'Bang tinh gia.xlsx'

# Load workbook với công thức
wb = load_workbook(file_path, data_only=False)
wb_data = load_workbook(file_path, data_only=True)

print("="*80)
print("PHAN TICH CHI TIET FILE EXCEL")
print("="*80)

sheet_name = 'BG - Hop Song Nap Cai Pizza'
ws = wb[sheet_name]
ws_data = wb_data[sheet_name]

# Đọc các dòng quan trọng với giá trị thực tế
print("\n=== SHEET CHINH: THONG SO NAP (Cot A-P, Row 1-30) ===\n")
for row_idx in range(1, 31):
    row_values = []
    has_data = False
    for col_idx in range(1, 17):  # A-P
        cell = ws_data.cell(row=row_idx, column=col_idx)
        if cell.value is not None:
            has_data = True
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            row_values.append(f"{col_letter}{row_idx}: {cell.value}")
    
    if has_data:
        print(f"Row {row_idx}:")
        for val in row_values:
            print(f"  {val}")
        print()

print("\n" + "="*80)
print("PHAN TICH CAC CONG THUC TRONG EXCEL")
print("="*80)

# Tìm các ô có công thức quan trọng
print("\n=== CONG THUC TINH TOAN (Row 1-50, Col A-P) ===\n")
for row_idx in range(1, 51):
    for col_idx in range(1, 17):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell_data = ws_data.cell(row=row_idx, column=col_idx)
        
        if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
            coord = cell.coordinate
            print(f"{coord}:")
            print(f"  Formula: {cell.value}")
            print(f"  Value: {cell_data.value}")
            print()

# Phân tích ô D4 (Đơn giá) - ô quan trọng nhất
print("\n" + "="*80)
print("PHAN TICH O DON GIA (D4)")
print("="*80)
cell_d4 = ws['D4']
cell_d4_data = ws_data['D4']
print(f"Formula D4: {cell_d4.value}")
print(f"Value D4: {cell_d4_data.value}")

# Phân tích sheet tổng hợp
print("\n" + "="*80)
print("BANG TINH GIA TONG HOP")
print("="*80)

ws2 = wb_data['Bảng tính giá Tổng hợp']
for row_idx in range(1, 35):
    row_values = []
    has_data = False
    for col_idx in range(1, 16):
        cell = ws2.cell(row=row_idx, column=col_idx)
        if cell.value is not None:
            has_data = True
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            row_values.append(f"{col_letter}: {cell.value}")
    
    if has_data:
        print(f"\nRow {row_idx}:")
        for val in row_values[:10]:  # Chi in 10 cot dau
            print(f"  {val}")
