import pandas as pd
import openpyxl

# Đọc file Excel
file_path = 'Bang tinh gia.xlsx'

# Load workbook để xem cấu trúc
wb = openpyxl.load_workbook(file_path, data_only=True)
print("Sheet names:", wb.sheetnames)
print("\n" + "="*50 + "\n")

# Đọc từng sheet
for sheet_name in wb.sheetnames:
    print(f"\n--- Sheet: {sheet_name} ---")
    ws = wb[sheet_name]
    
    # In ra một số dòng đầu
    print(f"Dimensions: {ws.dimensions}")
    print("\nFirst 20 rows:")
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if i <= 20:
            print(f"Row {i}: {row}")
    
    # Thử đọc với pandas
    try:
        df = pd.read_excel(file_path, sheet_name=sheet_name)
        print(f"\nPandas DataFrame shape: {df.shape}")
        print("\nDataFrame head:")
        print(df.head(15))
        print("\nDataFrame columns:")
        print(df.columns.tolist())
    except Exception as e:
        print(f"Error reading with pandas: {e}")
    
    print("\n" + "="*50 + "\n")
