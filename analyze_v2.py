# -*- coding: utf-8 -*-
import sys
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import pandas as pd
import openpyxl
from openpyxl import load_workbook

file_path = 'Bang tinh gia.xlsx'

# Load workbook
wb = load_workbook(file_path, data_only=False)
wb_data = load_workbook(file_path, data_only=True)

sheet_name = 'BG - Hộp Sóng Nắp Cài Pizza'
ws = wb[sheet_name]
ws_data = wb_data[sheet_name]

print("="*80)
print("PHAN TICH FILE EXCEL - SHEET NAP")
print("="*80)

# Đọc các giá trị quan trọng
print("\n=== CAC GIA TRI QUAN TRONG (30 dong dau) ===\n")
for row_idx in range(1, 31):
    row_data = []
    for col_idx in range(1, 20):  # A-S
        cell = ws_data.cell(row=row_idx, column=col_idx)
        if cell.value is not None:
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            row_data.append(f"{col_letter}: {cell.value}")
    
    if row_data:
        print(f"Row {row_idx}: " + " | ".join(row_data[:8]))

# Tìm công thức tính đơn giá (thường ở D4)
print("\n" + "="*80)
print("CONG THUC TINH DON GIA")
print("="*80)

for row_idx in range(1, 20):
    for col_idx in range(1, 10):
        cell = ws.cell(row=row_idx, column=col_idx)
        if cell.value and isinstance(cell.value, str) and '=' in str(cell.value):
            coord = cell.coordinate
            cell_value = ws_data.cell(row=row_idx, column=col_idx).value
            print(f"\n{coord}:")
            print(f"  Formula: {cell.value}")
            print(f"  Result: {cell_value}")

# Phân tích cấu trúc chi tiết
print("\n" + "="*80)
print("CAU TRUC CHI TIET (20 dong dau, 15 cot dau)")
print("="*80)

for row_idx in range(1, 21):
    print(f"\nRow {row_idx}:")
    for col_idx in range(1, 16):
        cell = ws_data.cell(row=row_idx, column=col_idx)
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        if cell.value is not None:
            print(f"  {col_letter}{row_idx} = {cell.value}")
