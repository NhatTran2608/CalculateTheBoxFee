import pandas as pd
import openpyxl
from openpyxl import load_workbook

# Đọc file Excel chi tiết để hiểu công thức
file_path = 'Bang tinh gia.xlsx'

# Load workbook với công thức
wb = load_workbook(file_path, data_only=False)
wb_data = load_workbook(file_path, data_only=True)

print("="*80)
print("PHÂN TÍCH CHI TIẾT FILE EXCEL - SHEET 'BG - Hộp Sóng Nắp Cài Pizza'")
print("="*80)

sheet_name = 'BG - Hộp Sóng Nắp Cài Pizza'
ws = wb[sheet_name]
ws_data = wb_data[sheet_name]

# In ra các ô quan trọng với công thức
print("\n--- CÁC Ô QUAN TRỌNG VỚI CÔNG THỨC ---\n")

# Duyệt qua các dòng quan trọng
for row in range(1, 50):
    for col in range(1, 20):
        cell = ws.cell(row=row, column=col)
        cell_data = ws_data.cell(row=row, column=col)
        
        # Chỉ in các ô có công thức hoặc giá trị quan trọng
        if cell.value and (isinstance(cell.value, str) and cell.value.startswith('=')):
            coord = cell.coordinate
            print(f"{coord}: Formula = {cell.value}")
            print(f"       Value = {cell_data.value}")
            print()

print("\n" + "="*80)
print("PHÂN TÍCH BẢNG TỔNG HỢP")
print("="*80)

sheet_name2 = 'Bảng tính giá Tổng hợp'
ws2 = wb[sheet_name2]
ws2_data = wb_data[sheet_name2]

print("\n--- CÁC Ô CÔNG THỨC TRONG BẢNG TỔNG HỢP ---\n")

for row in range(1, 35):
    for col in range(1, 16):
        cell = ws2.cell(row=row, column=col)
        cell_data = ws2_data.cell(row=row, column=col)
        
        if cell.value and (isinstance(cell.value, str) and cell.value.startswith('=')):
            coord = cell.coordinate
            print(f"{coord}: Formula = {cell.value}")
            print(f"       Value = {cell_data.value}")
            print()

# Đọc chi tiết cấu trúc dữ liệu
print("\n" + "="*80)
print("CẤU TRÚC DỮ LIỆU SHEET CHÍNH")
print("="*80)

ws_main = wb_data['BG - Hộp Sóng Nắp Cài Pizza']

# In 50 dòng đầu tiên với tất cả giá trị
for row_idx in range(1, 51):
    row_data = []
    for col_idx in range(1, 47):  # Đến cột AT (46)
        cell_value = ws_main.cell(row=row_idx, column=col_idx).value
        if cell_value is not None:
            row_data.append(f"{openpyxl.utils.get_column_letter(col_idx)}: {cell_value}")
    
    if row_data:
        print(f"\nRow {row_idx}:")
        for item in row_data:
            print(f"  {item}")
