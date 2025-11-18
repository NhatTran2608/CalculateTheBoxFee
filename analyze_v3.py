# -*- coding: utf-8 -*-
import sys
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl
from openpyxl import load_workbook

file_path = 'Bang tinh gia.xlsx'
wb = load_workbook(file_path, data_only=False)
wb_data = load_workbook(file_path, data_only=True)

sheet_name = 'BG - Hộp Sóng Nắp Cài Pizza'
ws = wb[sheet_name]
ws_data = wb_data[sheet_name]

print("="*80)
print("TIM CONG THUC TINH TONG CHI PHI (O L72)")
print("="*80)

# Xem ô L72
cell_l72 = ws['L72']
cell_l72_data = ws_data['L72']
print(f"\nL72 - TONG CHI PHI:")
print(f"  Formula: {cell_l72.value}")
print(f"  Value: {cell_l72_data.value}")

# Tìm các ô từ L60-L80
print("\n" + "="*80)
print("CAC O CHI PHI QUAN TRONG (L60-L80)")
print("="*80)

for row_idx in range(60, 81):
    cell = ws[f'L{row_idx}']
    cell_data = ws_data[f'L{row_idx}']
    cell_a = ws_data[f'A{row_idx}']
    
    if cell.value or cell_a.value:
        print(f"\nL{row_idx} ({cell_a.value}):")
        if cell.value and isinstance(cell.value, str) and '=' in str(cell.value):
            print(f"  Formula: {cell.value}")
        print(f"  Value: {cell_data.value}")

# Xem thêm các ô liên quan đến tính toán
print("\n" + "="*80)
print("CAC O TINH TOAN KHAC (Row 25-80, Col A,B,L)")
print("="*80)

for row_idx in range(25, 81):
    cell_a = ws_data[f'A{row_idx}']
    cell_b = ws_data[f'B{row_idx}']
    cell_l_formula = ws[f'L{row_idx}']
    cell_l = ws_data[f'L{row_idx}']
    
    if cell_a.value or cell_l.value:
        print(f"\nRow {row_idx}:")
        if cell_a.value:
            print(f"  A: {cell_a.value}")
        if cell_b.value:
            print(f"  B: {cell_b.value}")
        if cell_l_formula.value and isinstance(cell_l_formula.value, str) and '=' in str(cell_l_formula.value):
            print(f"  L Formula: {cell_l_formula.value[:100]}...")  # Chi in 100 ky tu dau
        if cell_l.value:
            print(f"  L Value: {cell_l.value}")

# Xem thêm xả lô
print("\n" + "="*80)
print("TINH TOAN XA LO")
print("="*80)

for row_idx in range(23, 30):
    for col_letter in ['A', 'B', 'C', 'D', 'E', 'F']:
        cell = ws_data[f'{col_letter}{row_idx}']
        cell_formula = ws[f'{col_letter}{row_idx}']
        if cell.value is not None:
            print(f"{col_letter}{row_idx}: {cell.value}", end=" | ")
            if cell_formula.value and isinstance(cell_formula.value, str) and '=' in str(cell_formula.value):
                print(f"Formula: {cell_formula.value[:50]}", end="")
        print()
