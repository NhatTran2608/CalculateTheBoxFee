# -*- coding: utf-8 -*-
import sys
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl
from openpyxl import load_workbook

file_path = 'Bang tinh gia.xlsx'
wb = load_workbook(file_path, data_only=False)
wb_data = load_workbook(file_path, data_only=True)

sheet_name = 'BG - Hộp Sóng Nắp Cài Pizza'
ws = wb[sheet_name]
ws_data = wb_data[sheet_name]

print("="*100)
print("CHI TIET CAC KHOAN CHI PHI (L88-L154)")
print("="*100)

for row_idx in range(88, 155):
    cell_a = ws_data[f'A{row_idx}']
    cell_b = ws_data[f'B{row_idx}']
    cell_l_formula = ws[f'L{row_idx}']
    cell_l = ws_data[f'L{row_idx}']
    
    # Chi in nhung dong co gia tri
    if cell_a.value or cell_l.value:
        print(f"\nRow {row_idx} - {cell_a.value}:")
        if cell_b.value is not None:
            print(f"  Column B: {cell_b.value}")
        
        if cell_l_formula.value and isinstance(cell_l_formula.value, str) and '=' in str(cell_l_formula.value):
            formula_str = str(cell_l_formula.value)
            if len(formula_str) > 150:
                print(f"  Formula: {formula_str[:150]}...")
            else:
                print(f"  Formula: {formula_str}")
        
        if cell_l.value is not None:
            print(f"  Value: {cell_l.value}")

# Xem thêm row 23-25 (xả lô)
print("\n" + "="*100)
print("TINH XA LO (Row 23-26)")
print("="*100)

for row_idx in range(23, 27):
    print(f"\nRow {row_idx}:")
    for col_idx in range(1, 15):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        cell = ws_data[f'{col_letter}{row_idx}']
        cell_formula = ws[f'{col_letter}{row_idx}']
        
        if cell.value is not None:
            print(f"  {col_letter}: {cell.value}", end="")
            if cell_formula.value and isinstance(cell_formula.value, str) and '=' in str(cell_formula.value):
                print(f" [F: {str(cell_formula.value)[:50]}]", end="")
            print()

# Lãi suất
print("\n" + "="*100)
print("LAI SUAT (Row 71)")
print("="*100)
cell_i71 = ws_data['I71']
print(f"I71 (% lãi): {cell_i71.value}")
