# -*- coding: utf-8 -*-
import sys
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl
from openpyxl import load_workbook

file_path = 'Bang tinh gia.xlsx'
wb = load_workbook(file_path, data_only=False)
wb_data = load_workbook(file_path, data_only=True)

sheet_name = 'BG - Hộp Sóng Nắp Cài Pizza'
ws = wb[sheet_name]
ws_data = wb_data[sheet_name]

print("="*100)
print("TINH XA LO VA SO TO")
print("="*100)

# Xem các ô D, E, G từ row 88-98 (chứa diện tích và số tờ)
for row_idx in range(88, 99):
    cell_a = ws_data[f'A{row_idx}']
    cell_d_formula = ws[f'D{row_idx}']
    cell_d = ws_data[f'D{row_idx}']
    cell_e_formula = ws[f'E{row_idx}']
    cell_e = ws_data[f'E{row_idx}']
    cell_g_formula = ws[f'G{row_idx}']
    cell_g = ws_data[f'G{row_idx}']
    
    print(f"\nRow {row_idx} - {cell_a.value}:")
    
    if cell_d_formula.value and isinstance(cell_d_formula.value, str) and '=' in str(cell_d_formula.value):
        print(f"  D (Dai): {str(cell_d_formula.value)[:80]}")
    if cell_d.value:
        print(f"  D Value: {cell_d.value}")
    
    if cell_e_formula.value and isinstance(cell_e_formula.value, str) and '=' in str(cell_e_formula.value):
        print(f"  E (Rong): {str(cell_e_formula.value)[:80]}")
    if cell_e.value:
        print(f"  E Value: {cell_e.value}")
    
    if cell_g_formula.value and isinstance(cell_g_formula.value, str) and '=' in str(cell_g_formula.value):
        print(f"  G (So to): {str(cell_g_formula.value)[:80]}")
    if cell_g.value:
        print(f"  G Value: {cell_g.value}")

# Xem các tham số giá trong row 9-10
print("\n" + "="*100)
print("CAC THAM SO GIA (Row 9-10, Col I-O)")
print("="*100)

for row_idx in [9, 10]:
    print(f"\nRow {row_idx}:")
    for col_letter in ['I', 'J', 'K', 'L', 'M', 'N', 'O']:
        cell = ws_data[f'{col_letter}{row_idx}']
        if cell.value is not None:
            print(f"  {col_letter}: {cell.value}")

# Xem thêm về số màu và máy in
print("\n" + "="*100)
print("THONG SO IN AN (Row 6-7, Col K-N)")
print("="*100)

for row_idx in [6, 7]:
    print(f"\nRow {row_idx}:")
    for col_letter in ['K', 'L', 'M', 'N']:
        cell = ws_data[f'{col_letter}{row_idx}']
        if cell.value is not None:
            print(f"  {col_letter}: {cell.value}")
